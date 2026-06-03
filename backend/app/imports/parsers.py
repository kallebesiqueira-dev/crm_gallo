"""Read an uploaded CSV / XLSX file into a uniform `(headers, rows)` shape.

Both parsers return:
  * `headers`  — the first row, as the raw (un-normalised) strings the
    user typed. Header→field mapping happens in `spec.py`.
  * `rows`     — a list of `dict[str, str]` keyed by raw header. Every
    value is coerced to a stripped string (numbers from XLSX become
    their string form) so the per-row Pydantic validator sees one
    consistent input type regardless of source format.

Design notes:
  * CSV is decoded `utf-8-sig` so a BOM from Excel-exported CSVs is
    stripped from the first header. `csv.Sniffer` picks `,` vs `;`
    (German/Swiss Excel defaults to `;`).
  * XLSX uses openpyxl `read_only=True` so a large sheet streams row by
    row instead of loading the whole workbook into memory.
  * `MAX_ROWS` caps the file so one upload can't pin the worker for
    minutes or OOM it. Rows beyond the cap raise `ImportParseError`
    (a whole-file failure — the job goes `failed`, nothing imported).
"""

from __future__ import annotations

import csv
import io

# Hard cap on data rows (excludes the header). A 50k-row contact import
# is already an unusually large single upload for an SMB CRM; beyond
# this we'd want chunked/resumable uploads, tracked as a follow-up.
MAX_ROWS = 50_000

_CSV_CONTENT_TYPES = {
    "text/csv",
    "application/csv",
    "text/plain",
    "application/vnd.ms-excel",  # some browsers label .csv this way
}
_XLSX_CONTENT_TYPES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/octet-stream",  # fallback — disambiguated by extension
}


class ImportParseError(Exception):
    """Whole-file parse failure — unreadable, empty, or over the row cap."""


def _is_xlsx(content_type: str, filename: str) -> bool:
    name = (filename or "").lower()
    if name.endswith(".xlsx"):
        return True
    if name.endswith(".csv"):
        return False
    return (content_type or "") in _XLSX_CONTENT_TYPES


def parse_table(
    content: bytes, content_type: str, filename: str
) -> tuple[list[str], list[dict[str, str]]]:
    """Dispatch on extension/content-type and return `(headers, rows)`."""
    if _is_xlsx(content_type, filename):
        return _parse_xlsx(content)
    return _parse_csv(content)


def _clean_headers(headers: list[str]) -> list[str]:
    cleaned = [(h or "").strip() for h in headers]
    if not any(cleaned):
        raise ImportParseError("No column headers found in the first row.")
    return cleaned


def _parse_csv(content: bytes) -> tuple[list[str], list[dict[str, str]]]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise ImportParseError("File is not valid UTF-8 text.") from exc

    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    except csv.Error:
        # Single-column files (no delimiter to sniff) — default to comma.
        dialect = csv.excel

    reader = csv.reader(io.StringIO(text), dialect)
    try:
        first = next(reader)
    except StopIteration as exc:
        raise ImportParseError("File is empty.") from exc

    headers = _clean_headers(first)
    rows: list[dict[str, str]] = []
    for raw in reader:
        # Skip fully-blank lines (trailing newline, separator rows).
        if not any((cell or "").strip() for cell in raw):
            continue
        if len(rows) >= MAX_ROWS:
            raise ImportParseError(f"File exceeds the {MAX_ROWS:,}-row limit.")
        rows.append(
            {headers[i]: (raw[i].strip() if i < len(raw) else "") for i in range(len(headers))}
        )
    return headers, rows


def _parse_xlsx(content: bytes) -> tuple[list[str], list[dict[str, str]]]:
    # Imported lazily so the module stays cheap to import and CSV-only
    # deployments don't pay for openpyxl at load time.
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dep is in requirements
        raise ImportParseError("XLSX support is not available on this server.") from exc

    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:  # openpyxl raises a grab-bag of errors
        raise ImportParseError("File is not a readable .xlsx workbook.") from exc

    ws = wb.active
    if ws is None:
        raise ImportParseError("Workbook has no active sheet.")

    row_iter = ws.iter_rows(values_only=True)
    try:
        first = next(row_iter)
    except StopIteration as exc:
        raise ImportParseError("File is empty.") from exc

    headers = _clean_headers([_cell_to_str(c) for c in first])
    rows: list[dict[str, str]] = []
    for raw in row_iter:
        values = [_cell_to_str(c) for c in raw]
        if not any(values):
            continue
        if len(rows) >= MAX_ROWS:
            wb.close()
            raise ImportParseError(f"File exceeds the {MAX_ROWS:,}-row limit.")
        rows.append(
            {headers[i]: (values[i] if i < len(values) else "") for i in range(len(headers))}
        )
    wb.close()
    return headers, rows


def _cell_to_str(value: object) -> str:
    """Coerce an XLSX cell to the same stripped-string shape as CSV.

    openpyxl yields native types (int/float/datetime/None). We render
    them as strings so the row validator's coercion path is identical
    for both formats. Floats that are whole numbers (Excel stores every
    number as a float) render without the `.0` tail so `"2"` not `"2.0"`.
    """
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()
